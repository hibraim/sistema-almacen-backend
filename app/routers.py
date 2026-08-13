from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime
from typing import Optional, List, Dict
import uuid
import shutil
import os
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from .database import get_db
from .models import Producto, Area, SubArea, Categoria, Entrada, Salida

router = APIRouter(prefix="/api/v1")

UPLOAD_DIR = "static/uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# --- CATEGORÍAS ---
@router.get("/categorias")
def listar_categorias(db: Session = Depends(get_db)):
    cats = db.query(Categoria).all()
    if not cats:
        cat_def = Categoria(nombre="Material General", descripcion="Insumos municipales")
        db.add(cat_def)
        db.commit()
        cats = db.query(Categoria).all()
    return {"data": [{"id": c.id, "nombre": c.nombre, "descripcion": c.descripcion} for c in cats]}

@router.post("/categorias")
def crear_categoria(data: dict, db: Session = Depends(get_db)):
    nombre_cat = data.get("nombre")
    if not nombre_cat:
        raise HTTPException(status_code=400, detail="El nombre es obligatorio")
    nueva = Categoria(nombre=nombre_cat, descripcion=data.get("descripcion", "General"))
    db.add(nueva)
    db.commit()
    db.refresh(nueva)
    return {"message": "Categoría creada", "data": {"id": nueva.id, "nombre": nueva.nombre}}

@router.delete("/categorias/{cat_id}")
def eliminar_categoria(cat_id: int, db: Session = Depends(get_db)):
    cat = db.query(Categoria).filter(Categoria.id == cat_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="No encontrada")
    db.delete(cat)
    db.commit()
    return {"message": "Eliminada"}


# --- ÁREAS Y DIRECCIONES ---
@router.get("/areas")
def listar_areas(db: Session = Depends(get_db)):
    areas = db.query(Area).all()
    if not areas:
        area_def = Area(nombre="Dirección de Alumbrado Público", encargado="Ing. Municipal", cargo="Director")
        db.add(area_def)
        db.commit()
        areas = db.query(Area).all()
        
    resultado = []
    for a in areas:
        subs = db.query(SubArea).filter(SubArea.area_id == a.id).all()
        resultado.append({
            "id": a.id,
            "nombre": a.nombre,
            "encargado": a.encargado,
            "cargo": a.cargo,
            "subareas": [{"id": s.id, "nombre": s.nombre, "encargado": s.encargado, "cargo": s.cargo} for s in subs]
        })
    return {"data": resultado}

@router.post("/areas")
def crear_area(data: dict, db: Session = Depends(get_db)):
    nueva = Area(
        nombre=data.get("nombre"),
        encargado=data.get("encargado", "Sin asignar"),
        cargo=data.get("cargo", "Director")
    )
    db.add(nueva)
    db.commit()
    return {"message": "Área creada"}

@router.delete("/areas/{area_id}")
def eliminar_area(area_id: int, db: Session = Depends(get_db)):
    area = db.query(Area).filter(Area.id == area_id).first()
    if not area:
        raise HTTPException(status_code=404, detail="No encontrada")
    db.delete(area)
    db.commit()
    return {"message": "Eliminada"}

@router.post("/subareas")
def crear_subarea(data: dict, db: Session = Depends(get_db)):
    nueva = SubArea(
        area_id=data.get("area_id"),
        nombre=data.get("nombre"),
        encargado=data.get("encargado", "Sin asignar"),
        cargo=data.get("cargo", "Jefe")
    )
    db.add(nueva)
    db.commit()
    return {"message": "Sub-área creada"}

@router.delete("/subareas/{sub_id}")
def eliminar_subarea(sub_id: int, db: Session = Depends(get_db)):
    sub = db.query(SubArea).filter(SubArea.id == sub_id).first()
    if not sub:
        raise HTTPException(status_code=404, detail="No encontrada")
    db.delete(sub)
    db.commit()
    return {"message": "Eliminada"}


# --- CATÁLOGO MAESTRO ---
@router.post("/articulos-catalogo")
async def crear_articulo_catalogo(
    codigo_interno: str = Form(...),
    nombre: str = Form(...),
    categoria_id: int = Form(1),
    unidad_medida: str = Form("Pieza"),
    descripcion: Optional[str] = Form(""),
    imagen: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    existe = db.query(Producto).filter(
        (Producto.codigo_interno == codigo_interno) | (Producto.nombre.ilike(nombre.strip()))
    ).first()
    
    if existe:
        raise HTTPException(status_code=400, detail="El código o nombre ya existe.")

    image_path = None
    if imagen and imagen.filename:
        file_extension = imagen.filename.split(".")[-1]
        file_name = f"{uuid.uuid4()}.{file_extension}"
        file_location = os.path.join(UPLOAD_DIR, file_name)
        with open(file_location, "wb+") as file_object:
            shutil.copyfileobj(imagen.file, file_object)
        image_path = f"/{UPLOAD_DIR}/{file_name}"

    nuevo = Producto(
        id=str(uuid.uuid4()),
        codigo_interno=codigo_interno,
        nombre=nombre.strip(),
        descripcion=descripcion or "",
        categoria_id=categoria_id,
        stock_actual=0,
        stock_minimo=5,
        stock_maximo=1000,
        unidad_medida=unidad_medida,
        imagen_principal=image_path
    )
    db.add(nuevo)
    db.commit()
    return {"message": "Artículo agregado al catálogo"}


# --- PRODUCTOS Y ENTRADAS ---
@router.get("/productos")
def listar_productos(db: Session = Depends(get_db)):
    prods = db.query(Producto).all()
    resultado = []
    for p in prods:
        cat = db.query(Categoria).filter(Categoria.id == p.categoria_id).first()
        resultado.append({
            "id": p.id,
            "codigo_interno": p.codigo_interno,
            "nombre": p.nombre,
            "descripcion": p.descripcion,
            "categoria_id": p.categoria_id,
            "categoria_nombre": cat.nombre if cat else "General",
            "stock_actual": p.stock_actual,
            "stock_minimo": p.stock_minimo,
            "stock_maximo": p.stock_maximo,
            "unidad_medida": p.unidad_medida,
            "imagen_principal": p.imagen_principal
        })
    return {"data": resultado}

@router.post("/productos")
async def crear_o_aumentar_producto(
    codigo_interno: str = Form(...),
    nombre: str = Form(...),
    descripcion: Optional[str] = Form(""),
    categoria_id: int = Form(1),
    stock_actual: int = Form(0),
    unidad_medida: Optional[str] = Form("Pieza"),
    area_id: Optional[str] = Form(None),
    proveedor: Optional[str] = Form("Proveedor General"),
    recibio: Optional[str] = Form("Almacén SPM"),
    imagen: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    try:
        image_path = None
        if imagen and imagen.filename:
            file_extension = imagen.filename.split(".")[-1]
            file_name = f"{uuid.uuid4()}.{file_extension}"
            file_location = os.path.join(UPLOAD_DIR, file_name)
            with open(file_location, "wb+") as file_object:
                shutil.copyfileobj(imagen.file, file_object)
            image_path = f"/{UPLOAD_DIR}/{file_name}"

        prov_final = proveedor if proveedor and proveedor.strip() != "" else "Proveedor General"
        unidad_final = unidad_medida if unidad_medida and unidad_medida.strip() != "" else "Pieza"

        producto_existente = db.query(Producto).filter(
            (Producto.codigo_interno == codigo_interno) | (Producto.nombre.ilike(nombre.strip()))
        ).first()

        if producto_existente:
            producto_existente.stock_actual += stock_actual
            producto_existente.unidad_medida = unidad_final
            if image_path:
                producto_existente.imagen_principal = image_path
            db.commit()
            prod_id = producto_existente.id
        else:
            prod_id = str(uuid.uuid4())
            nuevo = Producto(
                id=prod_id,
                codigo_interno=codigo_interno,
                nombre=nombre.strip(),
                descripcion=descripcion or "",
                categoria_id=int(categoria_id),
                stock_actual=stock_actual,
                stock_minimo=5,
                stock_maximo=1000,
                unidad_medida=unidad_final,
                imagen_principal=image_path
            )
            db.add(nuevo)
            db.commit()

        nueva_entrada = Entrada(
            id=str(uuid.uuid4()),
            folio=f"ENT-{uuid.uuid4().hex[:4].upper()}",
            fecha=datetime.now(),
            producto_id=prod_id,
            cantidad=stock_actual,
            proveedor=prov_final,
            area_id=int(area_id) if area_id and str(area_id).isdigit() else None,
            recibio_nombre=recibio or "Almacén SPM",
            entrego_nombre=prov_final
        )
        db.add(nueva_entrada)
        db.commit()

        return {"message": "Entrada registrada correctamente"}
    except Exception as e:
        print("ERROR:", str(e))
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/productos/{prod_id}")
def actualizar_producto(prod_id: str, data: dict, db: Session = Depends(get_db)):
    p = db.query(Producto).filter(Producto.id == prod_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="No encontrado")
    p.codigo_interno = data.get("codigo_interno", p.codigo_interno)
    p.nombre = data.get("nombre", p.nombre)
    p.stock_actual = int(data.get("stock_actual", p.stock_actual))
    p.unidad_medida = data.get("unidad_medida", p.unidad_medida)
    db.commit()
    return {"message": "Actualizado"}

@router.delete("/productos/{prod_id}")
def eliminar_producto(prod_id: str, db: Session = Depends(get_db)):
    p = db.query(Producto).filter(Producto.id == prod_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="No encontrado")
    db.delete(p)
    db.commit()
    return {"message": "Eliminado"}


# --- MOVIMIENTOS Y KÁRDEX ---
@router.get("/movimientos")
def listar_movimientos(db: Session = Depends(get_db)):
    salidas = db.query(Salida).order_by(Salida.fecha.desc()).all()
    entradas = db.query(Entrada).order_by(Entrada.fecha.desc()).all()
    
    resultado = []
    for s in salidas:
        p = db.query(Producto).filter(Producto.id == s.producto_id).first()
        a = db.query(Area).filter(Area.id == s.area_id).first() if s.area_id else None
        resultado.append({
            "id": s.id,
            "tipo": "SALIDA",
            "folio": s.folio,
            "fecha": s.fecha.strftime("%d/%m/%Y %H:%M:%S") if s.fecha else "",
            "producto": p.nombre if p else "Desconocido",
            "cantidad": s.cantidad,
            "area": a.nombre if a else "General",
            "recibio": s.recibio_nombre,
            "entrego": s.entrego_nombre
        })
        
    for e in entradas:
        p = db.query(Producto).filter(Producto.id == e.producto_id).first()
        a = db.query(Area).filter(Area.id == e.area_id).first() if e.area_id else None
        resultado.append({
            "id": e.id,
            "tipo": "ENTRADA",
            "folio": e.folio,
            "fecha": e.fecha.strftime("%d/%m/%Y %H:%M:%S") if e.fecha else "",
            "producto": p.nombre if p else "Desconocido",
            "cantidad": e.cantidad,
            "area": a.nombre if a else "Almacén General",
            "proveedor": e.proveedor,
            "recibio": e.recibio_nombre,
            "entrego": e.entrego_nombre
        })
    return {"data": resultado}

@router.post("/salidas")
def registrar_salida(data: dict, db: Session = Depends(get_db)):
    prod = db.query(Producto).filter(Producto.id == data.get("producto_id")).first()
    if not prod:
        raise HTTPException(status_code=404, detail="Artículo no encontrado")
    
    cant = int(data.get("cantidad", 1))
    if prod.stock_actual < cant:
        raise HTTPException(status_code=400, detail=f"Stock insuficiente. Disponible: {prod.stock_actual}")

    prod.stock_actual -= cant

    salida = Salida(
        id=str(uuid.uuid4()),
        folio=data.get("folio", f"SAL-{uuid.uuid4().hex[:4]}"),
        fecha=datetime.now(),
        producto_id=prod.id,
        cantidad=cant,
        area_id=data.get("area_id"),
        subarea_id=data.get("subarea_id"),
        descripcion=data.get("descripcion", ""),
        recibio_nombre=data.get("recibio_nombre", "Solicitante"),
        entrego_nombre=data.get("entrego_nombre", "Almacén")
    )
    db.add(salida)
    db.commit()
    return {"message": "Salida registrada"}

@router.delete("/movimientos/limpiar")
def limpiar_kardex(db: Session = Depends(get_db)):
    try:
        db.query(Salida).delete()
        db.query(Entrada).delete()
        db.commit()
        return {"message": "Kárdex y movimientos de prueba eliminados correctamente"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# --- REPORTES PDF Y EXCEL ---
@router.get("/reportes/excel")
def descargar_excel(area: Optional[str] = None, db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario SPM"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ["Código", "Artículo", "Categoría", "Stock Actual", "Unidad", "Descripción"]
    ws.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    prods = db.query(Producto).all()
    for p in prods:
        cat = db.query(Categoria).filter(Categoria.id == p.categoria_id).first()
        ws.append([
            p.codigo_interno,
            p.nombre,
            cat.nombre if cat else "General",
            p.stock_actual,
            p.unidad_medida,
            p.descripcion
        ])
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_inventario_spm.xlsx"}
    )

@router.get("/reportes/pdf")
def descargar_pdf(area: Optional[str] = None, db: Session = Depends(get_db)):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1F4E78'),
        alignment=1,
        spaceAfter=15
    )
    
    elements.append(Paragraph("SISTEMA SPM - REPORTE GENERAL DE INVENTARIO", title_style))
    elements.append(Spacer(1, 10))
    
    data = [["Código", "Artículo", "Categoría", "Stock Actual", "Unidad"]]
    prods = db.query(Producto).all()
    
    for p in prods:
        cat = db.query(Categoria).filter(Categoria.id == p.categoria_id).first()
        data.append([
            str(p.codigo_interno),
            str(p.nombre),
            str(cat.nombre if cat else "General"),
            str(p.stock_actual),
            str(p.unidad_medida)
        ])
        
    t = Table(data, colWidths=[90, 250, 150, 90, 90])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9F9F9')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD'))
    ]))
    
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=reporte_inventario_spm.pdf"}
    )